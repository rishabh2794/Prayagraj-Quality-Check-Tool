import streamlit as st
import pandas as pd
import json
import io
from openpyxl import Workbook
from openpyxl.styles import PatternFill
import os

# Streamlit page configuration
st.set_page_config(layout="wide")
st.title("A-PAG QC LOG Prayagraj")

with st.expander("🛠️ Standard Operating Procedure (SOP) for Quality Check Tool", expanded=False):
    st.markdown("""
    If you face issues while using the tool, follow these steps:  

    ✅ **1. Disable Firewall & Antivirus**  
    - Turn off your system’s firewall and temporarily disable Windows Defender/Antivirus.  

    ✅ **2. Check Network Settings**  
    - Switch to a *public network* if using a private one.  

    ✅ **3. Ensure Stable Internet**  
    - Verify your *internet speed* is stable.  

    ✅ **4. Check Server Status**  
    - Ensure the *dashboard server is operational*.  

    ✅ **5. Prevent Data Loss**  
    - **Enable "Save My Responses"** to avoid data loss.  
    - **Download responses** regularly, especially before breaks or shutting down your device.  

    For further assistance, contact **Rishabh Chaturvedi** at **rishabh.c@a-pag.org** or **9873695374**.
    """)

# File uploader
uploaded_file = st.file_uploader(
    "Upload your CSV or Excel file (the report Excel with metadata header rows)",
    type=["csv", "xlsx", "xls"]
)

# Feedback file location
FEEDBACK_FILE = "feedback_data_prayagraj.json"

# Ensure the feedback file exists
if not os.path.exists(FEEDBACK_FILE):
    with open(FEEDBACK_FILE, "w") as f:
        json.dump({}, f)

# Load previous feedback data into session state
if "feedback" not in st.session_state:
    try:
        with open(FEEDBACK_FILE, "r") as f:
            st.session_state.feedback = json.load(f)
    except Exception:
        st.session_state.feedback = {}

# Pagination setup
ROWS_PER_PAGE = 10
if "page" not in st.session_state:
    st.session_state.page = 0  # Start on page 0

# Disapproval reasons
disapproval_reasons = [
    "After Photo-Missing",
    "After Photo-Wrong/Blurry",
    "Incomplete Work/Work Not Started",
    "Image taken from wrong angle"
]

def detect_header_row_excel(path_or_buffer, max_scan_rows=20):
    """
    Scan the first max_scan_rows rows of the sheet (no header) to find a row
    that contains 'Complaint Number' (case-insensitive). Returns the row index
    to use as header (0-based). If not found, returns 5 (safe default seen in your file).
    """
    try:
        temp = pd.read_excel(path_or_buffer, header=None, nrows=max_scan_rows)
    except Exception:
        try:
            path_or_buffer.seek(0)
            temp = pd.read_excel(path_or_buffer, header=None, nrows=max_scan_rows)
        except Exception:
            return 5
    for i in range(min(max_scan_rows, len(temp))):
        row_vals = temp.iloc[i].astype(str).str.strip().tolist()
        if any(v.lower() == "complaint number" for v in row_vals if v):
            return i
    return 5

def read_input_file(uploaded):
    """
    Read uploaded file robustly:
    - If Excel: detect header row automatically by scanning for 'Complaint Number'.
    - If CSV: try normal read; if it looks like metadata, attempt fallback header row 5.
    Returns cleaned DataFrame with column names stripped.
    """
    try:
        uploaded.seek(0)
    except Exception:
        pass

    if uploaded.name.lower().endswith((".xls", ".xlsx")):
        header_row = detect_header_row_excel(uploaded)
        try:
            uploaded.seek(0)
        except Exception:
            pass
        df = pd.read_excel(uploaded, header=header_row)
    else:
        try:
            uploaded.seek(0)
        except Exception:
            pass
        df_try = pd.read_csv(uploaded, nrows=10)
        if any(str(c).lower().startswith("unnamed") for c in df_try.columns) or "from date" in map(str.lower, df_try.columns):
            try:
                uploaded.seek(0)
            except Exception:
                pass
            df = pd.read_csv(uploaded, header=5)
        else:
            uploaded.seek(0)
            df = pd.read_csv(uploaded)
    # Normalize column names: convert to string and strip whitespace
    df.columns = df.columns.astype(str).str.strip()
    return df

# Required columns (exact names after normalization)
REQUIRED_COLS = {
    "Complaint Number",
    "Zone",
    "Ward",
    "Complaint Sub type",
    "Address",
    "Surveyor Name",
    "Complaint Description",
    "Upload Documents",
    "Resolved Documents",
    "Registration Location"
}

if uploaded_file:
    try:
        df = read_input_file(uploaded_file)
    except Exception as e:
        st.error(f"Error reading the file: {e}")
        st.stop()

    # Quick preview for debugging: show the first row of headers (limited)
    st.write("Detected column names (first 40):", df.columns.tolist()[:40])

    # Check required columns after stripping
    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        st.error(
            "Your file is missing required columns (after header detection & stripping): "
            f"{missing}\n\n"
            "Tip: I tried to auto-detect the header row. If your file places column names on a different row,\n"
            "re-save the file so the header row is a single row of column names, or tell me which row number\n"
            "contains the real column names and I'll adapt the app."
        )
        st.stop()

    # Filters (only Zone, Ward, Complaint Sub type)
    selected_zone = st.selectbox("Filter by Zone", ["All"] + df["Zone"].dropna().unique().tolist())
    selected_ward = st.selectbox("Filter by Ward", ["All"] + df["Ward"].dropna().unique().tolist())
    selected_subtype = st.selectbox("Filter by Complaint Sub type", ["All"] + df["Complaint Sub type"].dropna().unique().tolist())

    # Apply filters
    filtered_df = df.copy()
    if selected_zone != "All":
        filtered_df = filtered_df[filtered_df["Zone"] == selected_zone]
    if selected_ward != "All":
        filtered_df = filtered_df[filtered_df["Ward"] == selected_ward]
    if selected_subtype != "All":
        filtered_df = filtered_df[filtered_df["Complaint Sub type"] == selected_subtype]

    # Reset pagination if filtered data is smaller than the current page
    total_pages = len(filtered_df) // ROWS_PER_PAGE + (len(filtered_df) % ROWS_PER_PAGE > 0)
    if total_pages == 0:
        total_pages = 1
    if st.session_state.page >= total_pages:
        st.session_state.page = 0

    # Paginate filtered results
    start_idx = st.session_state.page * ROWS_PER_PAGE
    end_idx = start_idx + ROWS_PER_PAGE
    df_page = filtered_df.iloc[start_idx:end_idx]

    # Live Review Summary
    status_counts = {
        "Status Yet to be Updated": 0,
        "Not Reviewed(Incorrect Before/Poor Identification)": 0,
        "Correct": 0,
        "Incorrect": 0
    }

    for pid in filtered_df["Complaint Number"].astype(str):
        current_status = st.session_state.feedback.get(str(pid), {}).get("Quality", "Status Yet to be Updated")
        status_counts[current_status] = status_counts.get(current_status, 0) + 1

    # Sidebar Summary
    with st.sidebar:
        st.subheader("Review Summary 📊")
        st.write(f" **Correct:** {status_counts['Correct']}")
        st.write(f"**Incorrect:** {status_counts['Incorrect']}")
        st.write(f" **Not Reviewed(Incorrect Before/Poor Identification):** {status_counts['Not Reviewed(Incorrect Before/Poor Identification)']}")
        st.write(f" **Status Yet to be Updated:** {status_counts['Status Yet to be Updated']}")

        total_correct = status_counts['Correct']
        total_incorrect = status_counts['Incorrect']
        total_not_reviewed = status_counts['Not Reviewed(Incorrect Before/Poor Identification)']
        total_status_not_updated = status_counts['Status Yet to be Updated']

        if total_correct + total_incorrect > 0:
            live_percentage = (total_correct * 100) / (total_correct + total_incorrect)
            st.write(f" **Current QC Status %:** {live_percentage:.2f}%")

        total_reviewed = total_correct + total_incorrect
        total_possible = total_reviewed + total_not_reviewed + total_status_not_updated

        if total_possible > 0:
            status_percentage = (total_reviewed * 100) / total_possible
            st.write(f"**Current Sample Size %** :{status_percentage:.2f}%")
        else:
            st.write("**Current Sample Size %** :0.00%")

        st.write(f"**Number of QC Done:** {total_correct + total_incorrect}")

    # Check if the page is empty
    if df_page.empty:
        st.error(f"Page {st.session_state.page + 1} is empty. Total filtered rows: {len(filtered_df)}")
    else:
        for _, row in df_page.iterrows():
            # Ensure complaint id is a string key
            complaint_id = str(row["Complaint Number"]).strip()
            subtype = row["Complaint Sub type"]
            pre_image = row["Upload Documents"]
            post_image = row["Resolved Documents"]
            zone = row["Zone"]
            ward = row["Ward"]
            address = row.get("Address", "")
            surveyor = row.get("Surveyor Name", "")
            description = row.get("Complaint Description", "")
            registration_location = row.get("Registration Location", "")

            st.subheader(f"Complaint Number: {complaint_id} | Sub type: {subtype}")
            st.text(f"Zone: {zone} | Ward: {ward} | Surveyor: {surveyor} | Address: {address}")
            st.write(f"**Complaint Description:** {description}")
            st.write(f"**Registration Location:** {registration_location}")

            col1, col2 = st.columns(2)
            # Before Photo
            with col1:
                if pd.notna(pre_image) and str(pre_image).strip() != "":
                    st.markdown(
                        f'<img src="{pre_image}" style="width:500px; height:400px; object-fit:cover; border:1px solid #ccc;">',
                        unsafe_allow_html=True
                    )
                else:
                    st.warning("No Pre Image Provided")

            # After Photo
            with col2:
                if pd.notna(post_image) and str(post_image).strip() != "":
                    st.markdown(
                        f'<img src="{post_image}" style="width:500px; height:400px; object-fit:cover; border:1px solid #ccc;">',
                        unsafe_allow_html=True
                    )
                else:
                    st.warning("No Post Image Provided")

            saved_status = st.session_state.feedback.get(complaint_id, {}).get("Quality", "Status Yet to be Updated")

            status = st.radio(
                label=f"Status for Complaint Number {complaint_id}",
                options=["Status Yet to be Updated", "Not Reviewed(Incorrect Before/Poor Identification)", "Correct", "Incorrect"],
                key=f"status_{complaint_id}",
                index=["Status Yet to be Updated", "Not Reviewed(Incorrect Before/Poor Identification)", "Correct", "Incorrect"].index(saved_status) if saved_status in ["Status Yet to be Updated", "Not Reviewed(Incorrect Before/Poor Identification)", "Correct", "Incorrect"] else 0
            )

            saved_reason = st.session_state.feedback.get(complaint_id, {}).get("comment", "")
            reason = ""
            if status == "Incorrect":
                reason = st.selectbox(
                    label=f"Reason for Disapproval (ID {complaint_id})",
                    options=disapproval_reasons,
                    key=f"reason_{complaint_id}",
                    index=disapproval_reasons.index(saved_reason) if saved_reason in disapproval_reasons else 0
                )

            st.session_state.feedback[complaint_id] = {
                "Quality": status,
                "comment": reason if status == "Incorrect" else ""
            }

        total_pages = len(filtered_df) // ROWS_PER_PAGE + (len(filtered_df) % ROWS_PER_PAGE > 0)
        if total_pages == 0:
            total_pages = 1
        st.markdown(f"**Page {st.session_state.page + 1} of {total_pages}**")

        col1, col2, col3 = st.columns([1, 6, 1])
        with col1:
            if st.session_state.page > 0:
                if st.button("⬅️ Previous Page"):
                    st.session_state.page -= 1
                    st.rerun()

        with col2:
            # Show a reasonable range of page buttons
            left = max(0, st.session_state.page - 3)
            right = min(total_pages, st.session_state.page + 10)
            page_numbers = list(range(left, right))
            pages = st.columns(len(page_numbers)) if page_numbers else [st.empty()]
            for i, p in enumerate(page_numbers):
                with pages[i]:
                    if st.button(f"**{p + 1}**", key=f"page_{p}"):
                        st.session_state.page = p
                        st.rerun()

        with col3:
            if st.session_state.page < total_pages - 1:
                if st.button("Next Page ➡️"):
                    st.session_state.page += 1
                    st.rerun()

        if st.button("Save My Responses"):
            with open(FEEDBACK_FILE, "w") as f:
                json.dump(st.session_state.feedback, f)
            st.success("Responses Saved!")

        def create_excel_download(df_in):
            wb = Workbook()
            ws = wb.active
            headers = df_in.columns.tolist()
            ws.append(headers + ["Review Status", "Comments"])

            total_cols = len(headers) + 2  # original columns + Review Status + Comments
            for row_index, (_, row_data) in enumerate(df_in.iterrows(), start=1):
                complaint_id = str(row_data["Complaint Number"]).strip()
                feedback_data = st.session_state.feedback.get(complaint_id, {})
                status = feedback_data.get("Quality", "Status Yet to be Updated")
                comment = feedback_data.get("comment", "")

                row_values = row_data.tolist() + [status, comment]
                ws.append(row_values)

                fill_color = {
                    "Correct": "00FF00",
                    "Incorrect": "FF0000",
                    "Not Reviewed(Incorrect Before/Poor Identification)": "FFFF00",
                }.get(status, "FFFFFF")

                # fill from col 1 to total_cols inclusive
                for col in range(1, total_cols + 1):
                    ws.cell(row=row_index + 1, column=col).fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            return buffer

        if st.button("Download Excel"):
            excel_buffer = create_excel_download(filtered_df)
            st.download_button("Download the Excel file", excel_buffer, "qc_log_prayagraj.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
